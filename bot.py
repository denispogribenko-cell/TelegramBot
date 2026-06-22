import os
import re
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters

TOKEN = "8876148935:AAGqNaJFpz8pfZKgir5XT3hCB5cXNvFrSok"

REPORTS_FOLDER = "reports"
CACHE = {}


def clean_phone(text):
    digits = re.sub(r"\D", "", str(text))

    if digits.startswith("380"):
        digits = digits[3:]

    if digits.startswith("0") and len(digits) == 10:
        digits = digits[1:]

    return digits


def format_money(value):
    try:
        value = float(value)
        real_value = value / 0.9

        if real_value.is_integer():
            return int(real_value)

        return round(real_value, 2)
    except:
        return "не удалось рассчитать"


def translate_status(status):
    status_text = str(status).strip().lower()

    if status_text == "active":
        return "Активный"

    if "closed" in status_text:
        return "Временно закрыт"

    return status


def load_reports():
    global CACHE
    CACHE = {}

    print("Загружаю Excel-файлы...")

    for filename in os.listdir(REPORTS_FOLDER):
        if not filename.endswith(".xlsx"):
            continue

        if filename.startswith("~$"):
            continue

        file_path = os.path.join(REPORTS_FOLDER, filename)

        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        month = str(sheet["A1"].value).replace("Звітний місяць:", "").strip()

        headers = {}
        for cell in sheet[2]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column

        needed = [
            "Особовий рахунок",
            "MSISDN",
            "Статус MSISDN",
            "Тарифний план",
            "Загальна сума"
        ]

        if any(name not in headers for name in needed):
            print(f"Пропущен файл {filename}: не хватает колонок")
            continue

        for row in range(3, sheet.max_row + 1):
            msisdn = sheet.cell(row=row, column=headers["MSISDN"]).value
            phone = clean_phone(msisdn)

            if len(phone) < 9:
                continue

            item = {
                "month": month,
                "phone": phone,
                "account": sheet.cell(row=row, column=headers["Особовий рахунок"]).value,
                "tariff": sheet.cell(row=row, column=headers["Тарифний план"]).value,
                "status": translate_status(sheet.cell(row=row, column=headers["Статус MSISDN"]).value),
                "expenses": format_money(sheet.cell(row=row, column=headers["Загальна сума"]).value)
            }

            if phone not in CACHE:
                CACHE[phone] = []

            CACHE[phone].append(item)

    print(f"Готово. Загружено номеров: {len(CACHE)}")


def find_phone(phone):
    return CACHE.get(phone, [])


def make_text(item, index, total):
    return (
        f"{item['month']}\n"
        f"{item['phone']}\n"
        f"Л/с: {item['account']}\n"
        f"Тариф: {item['tariff']}\n"
        f"Состояние номера: {item['status']}\n"
        f"Расходы : {item['expenses']} грн\n\n"
        f"Месяц {index + 1} из {total}"
    )


def make_keyboard(index, total, phone):
    buttons = []
    row = []

    if index > 0:
        row.append(InlineKeyboardButton("◀️ Назад", callback_data=f"month_{phone}_{index - 1}"))

    if index < total - 1:
        row.append(InlineKeyboardButton("Вперёд ▶️", callback_data=f"month_{phone}_{index + 1}"))

    if row:
        buttons.append(row)

    return InlineKeyboardMarkup(buttons)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Привет! Отправь номер телефона, и я найду его в Excel-отчётах."
    )


async def reload_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Обновляю данные из Excel...")
    load_reports()
    await update.message.reply_text("Готово. Данные обновлены.")


async def message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = clean_phone(update.message.text)

    if len(phone) < 9:
        await update.message.reply_text("Отправь номер телефона.")
        return

    results = find_phone(phone)

    if not results:
        await update.message.reply_text("Номер не найден.")
        return

    last_index = len(results) - 1

    await update.message.reply_text(
        make_text(results[last_index], last_index, len(results)),
        reply_markup=make_keyboard(last_index, len(results), phone)
    )


async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data.replace("month_", "")
    phone, index = data.rsplit("_", 1)
    index = int(index)

    results = find_phone(phone)

    if not results:
        await query.edit_message_text("Номер не найден.")
        return

    await query.edit_message_text(
    make_text(results[index], index, len(results)),
    reply_markup=make_keyboard(index, len(results), phone)
)


load_reports()

app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("reload", reload_command))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message))
app.add_handler(CallbackQueryHandler(button))

app.run_polling()